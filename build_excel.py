"""
ClearMetric Startup Runway Calculator — Premium Excel Template
Product for Gumroad ($14.99)

4 Sheets:
  1. Runway Calculator — inputs, key metrics, runway, default alive verdict
  2. Monthly Projection — 36 months of cash, revenue, expenses, net, cumulative
  3. Scenario Comparison — 3 scenarios (current, lean, aggressive growth)
  4. How To Use — instructions

Design: Coral/Red-Orange palette (#C0392B primary, #922B21 dark, #FADBD8 input)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
import os

# ============================================================
# DESIGN SYSTEM — Coral/Red-Orange
# ============================================================
CORAL = "C0392B"
DARK_CORAL = "922B21"
WHITE = "FFFFFF"
INPUT_CORAL = "FADBD8"
LIGHT_GRAY = "F5F6FA"
MED_GRAY = "D5D8DC"
DARK_GRAY = "5D6D7E"
GREEN = "27AE60"
LIGHT_GREEN = "EAFAF1"
RED = "E74C3C"
LIGHT_RED = "FDEDEC"
YELLOW = "F39C12"
LIGHT_YELLOW = "FEF9E7"
ACCENT = "E74C3C"
LIGHT_CORAL = "FDEBD0"

FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=WHITE)
FONT_SUBTITLE = Font(name="Calibri", size=12, color="FADBD8", italic=True)
FONT_SECTION = Font(name="Calibri", size=13, bold=True, color=WHITE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_LABEL = Font(name="Calibri", size=11, color="2C3E50")
FONT_INPUT = Font(name="Calibri", size=12, color=DARK_CORAL, bold=True)
FONT_VALUE = Font(name="Calibri", size=11, color="2C3E50")
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color=CORAL)
FONT_SMALL = Font(name="Calibri", size=9, color=DARK_GRAY, italic=True)
FONT_BIG = Font(name="Calibri", size=28, bold=True, color=WHITE)
FONT_BIG_LABEL = Font(name="Calibri", size=12, bold=True, color="FADBD8")
FONT_GREEN = Font(name="Calibri", size=11, bold=True, color=GREEN)
FONT_RED = Font(name="Calibri", size=11, bold=True, color=RED)
FONT_WHITE_BOLD = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_CTA = Font(name="Calibri", size=12, bold=True, color=CORAL)

FILL_CORAL = PatternFill(start_color=CORAL, end_color=CORAL, fill_type="solid")
FILL_DARK = PatternFill(start_color=DARK_CORAL, end_color=DARK_CORAL, fill_type="solid")
FILL_INPUT = PatternFill(start_color=INPUT_CORAL, end_color=INPUT_CORAL, fill_type="solid")
FILL_GRAY = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
FILL_LIGHT = PatternFill(start_color=LIGHT_CORAL, end_color=LIGHT_CORAL, fill_type="solid")
FILL_GREEN = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
FILL_RED = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
FILL_YELLOW = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")

THIN = Border(
    left=Side("thin", MED_GRAY), right=Side("thin", MED_GRAY),
    top=Side("thin", MED_GRAY), bottom=Side("thin", MED_GRAY),
)
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center")


def header_bar(ws, row, c1, c2, text, fill=None):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1, value=text)
    cell.font = FONT_SECTION
    cell.fill = fill or FILL_CORAL
    cell.alignment = ALIGN_C
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).fill = fill or FILL_CORAL
        ws.cell(row=row, column=c).border = THIN


def label_input(ws, row, lc, vc, label, value=None, fmt=None, hint=None):
    cl = ws.cell(row=row, column=lc, value=label)
    cl.font = FONT_LABEL
    cl.fill = FILL_GRAY
    cl.border = THIN
    cl.alignment = ALIGN_L
    cv = ws.cell(row=row, column=vc, value=value)
    cv.font = FONT_INPUT
    cv.fill = FILL_INPUT
    cv.border = THIN
    cv.alignment = ALIGN_R
    if fmt:
        cv.number_format = fmt
    if hint:
        ch = ws.cell(row=row, column=vc + 1, value=hint)
        ch.font = FONT_SMALL
        ch.alignment = ALIGN_L


def label_calc(ws, row, lc, vc, label, formula, fmt=None, bold=False):
    cl = ws.cell(row=row, column=lc, value=label)
    cl.font = FONT_LABEL
    cl.fill = FILL_GRAY
    cl.border = THIN
    cl.alignment = ALIGN_L
    cv = ws.cell(row=row, column=vc, value=formula)
    cv.font = FONT_BOLD if bold else FONT_VALUE
    cv.fill = FILL_WHITE
    cv.border = THIN
    cv.alignment = ALIGN_R
    if fmt:
        cv.number_format = fmt


def cols(ws, widths):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


# ============================================================
# SHEET 1: RUNWAY CALCULATOR
# ============================================================
def build_runway_calculator(ws):
    ws.title = "Runway Calculator"
    ws.sheet_properties.tabColor = CORAL
    cols(ws, {"A": 2, "B": 36, "C": 18, "D": 18, "E": 4, "F": 36, "G": 18, "H": 2})

    for r in range(1, 75):
        for c in range(1, 9):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    # Title
    for r in range(1, 4):
        for c in range(2, 8):
            ws.cell(row=r, column=c).fill = FILL_DARK
    ws.merge_cells("B1:G1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells("B2:G2")
    ws.row_dimensions[2].height = 38
    title = ws.cell(row=2, column=2, value="STARTUP RUNWAY CALCULATOR")
    title.font = FONT_TITLE
    title.alignment = ALIGN_C
    ws.merge_cells("B3:G3")
    ws.row_dimensions[3].height = 22
    sub = ws.cell(row=3, column=2, value="How long does your money last? Plan fundraising. Model hires. Know when you'll break even.")
    sub.font = FONT_SUBTITLE
    sub.alignment = ALIGN_C

    # Left column: inputs
    header_bar(ws, 5, 2, 4, "CASH & REVENUE")
    label_input(ws, 6, 2, 3, "Current cash in bank", 500000, "$#,##0")
    label_input(ws, 7, 2, 3, "Monthly revenue", 10000, "$#,##0")
    label_input(ws, 8, 2, 3, "Monthly revenue growth rate", 0.10, "0.0%")

    header_bar(ws, 10, 2, 4, "MONTHLY BURN (by category)")
    label_input(ws, 11, 2, 3, "Salaries & payroll", 30000, "$#,##0")
    label_input(ws, 12, 2, 3, "Office/workspace", 3000, "$#,##0")
    label_input(ws, 13, 2, 3, "Software & tools", 2000, "$#,##0")
    label_input(ws, 14, 2, 3, "Marketing", 5000, "$#,##0")
    label_input(ws, 15, 2, 3, "Legal & accounting", 1500, "$#,##0")
    label_input(ws, 16, 2, 3, "Other expenses", 2000, "$#,##0")

    header_bar(ws, 18, 2, 4, "PLANNED HIRES")
    label_input(ws, 19, 2, 3, "Number of new hires", 2, "0")
    label_input(ws, 20, 2, 3, "Avg salary each ($/mo)", 6000, "$#,##0")
    label_input(ws, 21, 2, 3, "Hire start month", 4, "0")

    header_bar(ws, 23, 2, 4, "FUNDRAISING & ONE-TIME")
    label_input(ws, 24, 2, 3, "One-time expenses", 0, "$#,##0")
    label_input(ws, 25, 2, 3, "Target raise amount", 0, "$#,##0")
    label_input(ws, 26, 2, 3, "Expected close month", 0, "0")

    # Right column: results
    header_bar(ws, 5, 6, 7, "KEY METRICS", FILL_DARK)

    label_calc(ws, 6, 6, 7, "Total monthly burn", "=C11+C12+C13+C14+C15+C16", "$#,##0", bold=True)
    label_calc(ws, 7, 6, 7, "Net monthly burn", "=G6-C7", "$#,##0", bold=True)
    label_calc(ws, 8, 6, 7, "Gross margin", '=IF(C7>0,(C7-0)/C7,0)', "0.0%")
    label_calc(ws, 9, 6, 7, "Runway (months)", "=IF(G7>=0,\"N/A (profitable)\",C6/ABS(G7))", "0", bold=True)
    label_calc(ws, 10, 6, 7, "Break-even month", "=IF(G7>=0,1,\"See Monthly Projection\")", None)
    label_calc(ws, 11, 6, 7, "Start fundraising by (mo)", "=MAX(1,G9-6)", "0")
    label_calc(ws, 12, 6, 7, "Critical cash (3mo burn)", "=G6*3", "$#,##0")

    header_bar(ws, 14, 6, 7, "DEFAULT ALIVE?")
    ws.merge_cells("F15:G17")
    da_cell = ws.cell(row=15, column=6)
    da_cell.value = '=IF(C7*(1+C8)^11>=G6+(IF(C21<=12,C19*C20,0)),"✅ Default Alive","❌ Default Dead")'
    da_cell.font = Font(name="Calibri", size=14, bold=True, color=CORAL)
    da_cell.alignment = ALIGN_C
    for r in range(15, 18):
        for c in range(6, 8):
            ws.cell(row=r, column=c).fill = FILL_LIGHT
            ws.cell(row=r, column=c).border = THIN

    header_bar(ws, 19, 6, 7, "VERDICT")
    ws.merge_cells("F20:G22")
    verdict = ws.cell(row=20, column=6)
    verdict.value = '=IF(C7*(1+C8)^11>=G6+(IF(C21<=12,C19*C20,0)),"At current trends, revenue will exceed expenses by month 12. You\'re on a path to profitability.","Revenue won\'t cover expenses by month 12. Grow revenue faster, cut burn, or raise capital.")'
    verdict.font = FONT_SMALL
    verdict.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for r in range(20, 23):
        for c in range(6, 8):
            ws.cell(row=r, column=c).fill = FILL_WHITE
            ws.cell(row=r, column=c).border = THIN

    # Footer
    row = 25
    ws.merge_cells(f"B{row}:G{row}")
    ws.cell(row=row, column=2, value="© ClearMetric | clearmetric.gumroad.com | Startup Runway Calculator").font = FONT_SMALL
    ws.cell(row=row, column=2).alignment = ALIGN_C

    ws.protection.sheet = True
    ws.protection.set_password("")
    input_cells = [(6, 3), (7, 3), (8, 3), (11, 3), (12, 3), (13, 3), (14, 3), (15, 3), (16, 3),
                   (19, 3), (20, 3), (21, 3), (24, 3), (25, 3), (26, 3)]
    for r, c in input_cells:
        ws.cell(row=r, column=c).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 2: MONTHLY PROJECTION
# ============================================================
def build_monthly_projection(wb):
    ws = wb.create_sheet("Monthly Projection")
    ws.sheet_properties.tabColor = DARK_CORAL
    rc = "'Runway Calculator'"
    cols(ws, {"A": 2, "B": 6, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 2})

    for r in range(1, 4):
        for c in range(2, 8):
            ws.cell(row=r, column=c).fill = FILL_DARK
    ws.merge_cells("B1:G1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells("B2:G2")
    ws.row_dimensions[2].height = 38
    ws.cell(row=2, column=2, value="MONTHLY PROJECTION (36 months)").font = FONT_TITLE
    ws.cell(row=2, column=2).alignment = ALIGN_C
    ws.merge_cells("B3:G3")
    ws.cell(row=3, column=2, value="Cash, revenue, expenses, net, and cumulative. Based on Runway Calculator inputs.").font = FONT_SUBTITLE
    ws.cell(row=3, column=2).alignment = ALIGN_C

    headers = ["Month", "Cash", "Revenue", "Expenses", "Net", "Cumulative Net"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=5, column=2 + i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_CORAL
        cell.alignment = ALIGN_C
        cell.border = THIN

    burn_base = f"={rc}!C11+{rc}!C12+{rc}!C13+{rc}!C14+{rc}!C15+{rc}!C16"

    for m in range(1, 37):
        r = 5 + m
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=2, value=m).font = FONT_VALUE
        ws.cell(row=r, column=2).alignment = ALIGN_C
        ws.cell(row=r, column=2).border = THIN
        ws.cell(row=r, column=2).fill = FILL_GRAY

        if m == 1:
            rev_formula = f"={rc}!C7"
        else:
            rev_formula = f"=C{r-1}*(1+{rc}!C8)"
        exp_formula = f"{burn_base}+IF(B{r}>={rc}!C21,{rc}!C19*{rc}!C20,0)"

        ws.cell(row=r, column=3, value=rev_formula).font = FONT_VALUE
        ws.cell(row=r, column=3).number_format = "$#,##0"
        ws.cell(row=r, column=3).border = THIN

        ws.cell(row=r, column=4, value=exp_formula).font = FONT_VALUE
        ws.cell(row=r, column=4).number_format = "$#,##0"
        ws.cell(row=r, column=4).border = THIN

        ws.cell(row=r, column=5, value=f"=C{r}-D{r}").font = FONT_VALUE
        ws.cell(row=r, column=5).number_format = "$#,##0"
        ws.cell(row=r, column=5).border = THIN

        if m == 1:
            cash_formula = f"=MAX(0,{rc}!C6+E{r}-{rc}!C24+IF(B{r}={rc}!C26,{rc}!C25,0))"
        else:
            cash_formula = f"=MAX(0,F{r-1}+E{r}+IF(B{r}={rc}!C26,{rc}!C25,0))"
        ws.cell(row=r, column=6, value=cash_formula).font = FONT_BOLD
        ws.cell(row=r, column=6).number_format = "$#,##0"
        ws.cell(row=r, column=6).border = THIN

        if m == 1:
            cum_formula = f"=E{r}"
        else:
            cum_formula = f"=G{r-1}+E{r}"
        ws.cell(row=r, column=7, value=cum_formula).font = FONT_VALUE
        ws.cell(row=r, column=7).number_format = "$#,##0"
        ws.cell(row=r, column=7).border = THIN

        if m % 2 == 1:
            for c in range(3, 8):
                ws.cell(row=r, column=c).fill = FILL_GRAY


# ============================================================
# SHEET 3: SCENARIO COMPARISON
# ============================================================
def build_scenario_comparison(wb):
    ws = wb.create_sheet("Scenario Comparison")
    ws.sheet_properties.tabColor = "2E86C1"
    rc = "'Runway Calculator'"
    cols(ws, {"A": 2, "B": 28, "C": 16, "D": 4, "E": 16, "F": 4, "G": 16, "H": 2})

    for r in range(1, 55):
        for c in range(1, 9):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    for r in range(1, 4):
        for c in range(2, 8):
            ws.cell(row=r, column=c).fill = FILL_DARK
    ws.merge_cells("B1:G1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells("B2:G2")
    ws.row_dimensions[2].height = 38
    ws.cell(row=2, column=2, value="SCENARIO COMPARISON").font = FONT_TITLE
    ws.cell(row=2, column=2).alignment = ALIGN_C
    ws.merge_cells("B3:G3")
    ws.cell(row=3, column=2, value="Compare 3 paths: Current, Lean (cut burn), Aggressive growth. Green = inputs.").font = FONT_SUBTITLE
    ws.cell(row=3, column=2).alignment = ALIGN_C

    for c in range(2, 8):
        ws.cell(row=5, column=c).fill = FILL_CORAL
        ws.cell(row=5, column=c).border = THIN
    ws.cell(row=5, column=2, value="Parameter").font = FONT_WHITE_BOLD
    ws.cell(row=5, column=2).alignment = ALIGN_C
    ws.cell(row=5, column=3, value="Current").font = FONT_WHITE_BOLD
    ws.cell(row=5, column=3).alignment = ALIGN_C
    ws.cell(row=5, column=5, value="Lean").font = FONT_WHITE_BOLD
    ws.cell(row=5, column=5).alignment = ALIGN_C
    ws.cell(row=5, column=7, value="Aggressive").font = FONT_WHITE_BOLD
    ws.cell(row=5, column=7).alignment = ALIGN_C

    params = [
        (6, "Current cash", f"={rc}!C6", 500000, 500000, "$#,##0"),
        (7, "Monthly revenue", f"={rc}!C7", 10000, 10000, "$#,##0"),
        (8, "Revenue growth (mo)", f"={rc}!C8", 0.10, 0.10, "0.0%"),
        (9, "Total burn", f"={rc}!C11+{rc}!C12+{rc}!C13+{rc}!C14+{rc}!C15+{rc}!C16", 43500, 35000, "$#,##0"),
        (10, "Num hires", f"={rc}!C19", 2, 0, "0"),
        (11, "Salary per hire", f"={rc}!C20", 6000, 6000, "$#,##0"),
        (12, "Lean burn (edit)", "=C9", 35000, 35000, "$#,##0"),
        (13, "Aggressive growth (edit)", "=C8", 0.05, 0.15, "0.0%"),
    ]

    for r, label, val_a, val_b, val_c, fmt in params:
        ws.row_dimensions[r].height = 22
        ws.cell(row=r, column=2, value=label).font = FONT_LABEL
        ws.cell(row=r, column=2).fill = FILL_GRAY
        ws.cell(row=r, column=2).border = THIN

        ca = ws.cell(row=r, column=3, value=val_a)
        ca.font = FONT_BOLD
        ca.fill = FILL_LIGHT
        ca.border = THIN
        ca.alignment = ALIGN_C
        if fmt:
            ca.number_format = fmt

        cb = ws.cell(row=r, column=5, value=val_b)
        cb.font = FONT_INPUT
        cb.fill = FILL_INPUT
        cb.border = THIN
        cb.alignment = ALIGN_C
        if fmt:
            cb.number_format = fmt

        cc = ws.cell(row=r, column=7, value=val_c)
        cc.font = FONT_INPUT
        cc.fill = FILL_INPUT
        cc.border = THIN
        cc.alignment = ALIGN_C
        if fmt:
            cc.number_format = fmt

    header_bar(ws, 15, 2, 7, "RESULTS")

    def _result_row(r, label, formula_a, formula_b, formula_c, fmt):
        ws.row_dimensions[r].height = 22
        ws.cell(row=r, column=2, value=label).font = FONT_LABEL
        ws.cell(row=r, column=2).fill = FILL_GRAY
        ws.cell(row=r, column=2).border = THIN
        for col, formula in [(3, formula_a), (5, formula_b), (7, formula_c)]:
            cell = ws.cell(row=r, column=col, value=formula)
            cell.font = FONT_BOLD
            cell.fill = FILL_WHITE
            cell.border = THIN
            cell.alignment = ALIGN_C
            if fmt:
                cell.number_format = fmt

    _result_row(16, "Net burn", "=C9-C7", "=E9-E7", "=G9-G7", "$#,##0")
    _result_row(17, "Runway (months)", '=IF(C16>=0,"N/A",C6/ABS(C16))', '=IF(E16>=0,"N/A",E6/ABS(E16))', '=IF(G16>=0,"N/A",G6/ABS(G16))', "0")
    _result_row(18, "Break-even (approx)", "=IF(C16<=0,\"See Monthly\",\"Profitable\")", "=IF(E16<=0,\"See Monthly\",\"Profitable\")", "=IF(G16<=0,\"See Monthly\",\"Profitable\")", None)

    ws.protection.sheet = True
    ws.protection.set_password("")
    for r in range(6, 14):
        for c in [5, 7]:
            ws.cell(row=r, column=c).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 4: HOW TO USE
# ============================================================
def build_instructions(wb):
    ws = wb.create_sheet("How To Use")
    ws.sheet_properties.tabColor = DARK_GRAY
    cols(ws, {"A": 3, "B": 90})

    ws.merge_cells("A1:B2")
    c = ws.cell(row=1, column=1, value="HOW TO USE THE STARTUP RUNWAY CALCULATOR")
    c.font = FONT_TITLE
    c.fill = FILL_DARK
    c.alignment = ALIGN_C
    for r in range(1, 3):
        for co in range(1, 3):
            ws.cell(row=r, column=co).fill = FILL_DARK

    sections = [
        ("QUICK START", [
            "1. Open the 'Runway Calculator' tab and enter your numbers in the PINK/CORAL cells",
            "2. Key metrics appear instantly: runway, net burn, break-even, default alive verdict",
            "3. Use 'Monthly Projection' for a 36-month cash, revenue, expense, and net breakdown",
            "4. Use 'Scenario Comparison' to compare Current vs Lean vs Aggressive growth",
            "5. Start fundraising 6 months before your runway runs out",
        ]),
        ("WHAT IS RUNWAY?", [
            "Runway = how many months until your cash runs out at current burn and revenue",
            "Net burn = monthly expenses minus monthly revenue",
            "If revenue grows faster than expenses, you'll eventually break even",
            "Default Alive (Paul Graham): if trends continue, will you reach profitability?",
            "Default Dead: you'll run out of cash before reaching profitability — need to change course",
        ]),
        ("KEY METRICS EXPLAINED", [
            "Total monthly burn: sum of all expense categories",
            "Net monthly burn: burn minus revenue (negative = you're losing money each month)",
            "Break-even month: when revenue first equals or exceeds expenses",
            "Critical cash: 3 months of burn — below this, you're in danger zone",
            "Start fundraising by: 6 months before cash runs out — gives time to close a round",
        ]),
        ("SCENARIO COMPARISON TIPS", [
            "Current: uses your main Runway Calculator inputs",
            "Lean: reduce burn (e.g. cut marketing, defer hires) — edit the Lean burn cell",
            "Aggressive: higher revenue growth — edit the Aggressive growth cell",
            "Compare runway and break-even across scenarios to plan your strategy",
        ]),
        ("ABOUT THIS TEMPLATE", [
            "Version: 1.0 | Compatible with: Microsoft Excel 2016+, Google Sheets, LibreOffice Calc",
            "All projections assume revenue grows at your specified monthly rate",
            "This template is for startup planning only. Not financial advice.",
            "© 2026 ClearMetric. All Rights Reserved.",
            "Questions? Visit clearmetric.gumroad.com",
        ]),
    ]

    r = 4
    for title, items in sections:
        ws.cell(row=r, column=2, value=title).font = Font(name="Calibri", size=12, bold=True, color=CORAL)
        ws.cell(row=r, column=2).fill = FILL_LIGHT
        ws.cell(row=r, column=2).border = THIN
        r += 1
        for item in items:
            ws.cell(row=r, column=2, value=item).font = Font(name="Calibri", size=11, color="2C3E50")
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 22
            r += 1
        r += 1


# ============================================================
# MAIN
# ============================================================
def main():
    wb = openpyxl.Workbook()
    ws = wb.active

    print("Building Runway Calculator sheet...")
    build_runway_calculator(ws)

    print("Building Monthly Projection sheet...")
    build_monthly_projection(wb)

    print("Building Scenario Comparison sheet...")
    build_scenario_comparison(wb)

    print("Building How To Use sheet...")
    build_instructions(wb)

    wb.active = 0

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       "output", "ClearMetric-Startup-Runway-Calculator.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"Size: {os.path.getsize(out) / 1024:.1f} KB")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()

